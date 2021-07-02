import sys
from django.http.request import HttpRequest
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.http.response import HttpResponseRedirect, StreamingHttpResponse
import openpyxl
from .models import SelectedTool, Tool
from django.forms import forms
from decorator import decorator
from django.conf import settings
import datetime
from dateutil.relativedelta import relativedelta

import cv2
import threading
from pyzbar.pyzbar import decode
from pyzbar.pyzbar import ZBarSymbol

@decorator
def login_required(f, request, *args, **kwargs):
    """
    Makes sure the user connecting is known in the database.

    Args:
        f (function): the functio nto which this decorator is attached
        request: contains session variables, type of request, etc.

    Returns:
        f: calls the function to follow
        HttpResponseRedirect: redirect to login page if user is not connected.
    """
    name = request.session.get("name")
    if name is None:
        return HttpResponseRedirect('/login/')
    else :
        return f(request, *args, **kwargs)

@login_required
def index(request):
    cam = VideoCamera(request)
    """
    Search in 2 excel files infos about the connected user.

    Args:
        request: contains session variables, type of request, etc.

    Returns:
        render(function): Shows the index html page.
    """
    
    name = request.session.get("name")
    habilitations = settings.FILE_EXCEL_HABILITATION
    codes = settings.FILE_EXCEL_CODE_AFFAIRE
        
    context = {
        'name': name
    }

    wb = openpyxl.load_workbook(habilitations)
    ws = wb.active
    context['fullname'] = f"{request.user.first_name} {request.user.last_name.upper()}"
    for row in ws.iter_rows():
        if (row[0].value is not None and
            row[0].value.upper() == f"{request.user.last_name} {request.user.first_name}".upper()):
            
            dateElec = str(row[2].value)
            dateMedic = str(row[10].value)
            dateElec = datetime.datetime.strptime(dateElec, "%Y-%m-%d %H:%M:%S")
            dateMedic = datetime.datetime.strptime(dateMedic, '%Y-%m-%d %H:%M:%S')
            
            today = datetime.datetime(datetime.date.today().year, datetime.date.today().month, datetime.date.today().day) 
            if (dateElec + relativedelta(months=36) < today) :
                elecIsOld = True
            else :
                elecIsOld = False
            if (dateMedic + relativedelta(months=24) < today) :
                medicIsOld = True
            else :
                medicIsOld = False
            context['elecIsOld'] = elecIsOld
            context['medicIsOld'] = medicIsOld
            
            dateElec = datetime.datetime.strftime(dateElec, "%d/%m/%Y")
            dateMedic = datetime.datetime.strftime(dateMedic, "%d/%m/%Y")
            context['dateElec'] = dateElec
            context['dateMedic'] = dateMedic
            
            break
    
    wb = openpyxl.load_workbook(codes)
    ws = wb.active
    
    codes = {}
    row_nb = -1
    for i, row in enumerate(ws.iter_rows()):
        if (row[0].value is not None and
            row[0].value.upper() == request.user.last_name.upper()):
            row_nb = i
            break
    
    if row_nb != -1:
        for col in ws.iter_cols():
            if col[row_nb].value is not None:
                codes[col[4].value] = True
            else :
                codes[col[4].value] = False
    codes.pop(None, None)
    codes.pop('D = Devis\nM = Mesures\nR = Rapport', None)
    if codes is not None :
        context['codes'] = codes
    
    return render(request, 'main/index.html', context)

@login_required
def reserve(request, tool_id):
    return HttpResponse("Reserving tool {}".format(tool_id))

@login_required
def materiel(request):
    context = {
        'name': request.session.get("name")
    }
    return render(request, 'main/materiel.html', context)

@login_required
def consultation(request):
    return HttpResponse("consultation")

@login_required
def sortie(request):
    """
    Search in excel file for Categories, Tools, and infos.

    Args:
        request: contains session variables, type of request, etc.

    Returns:
        render(function): Shows the sortie html page.
    """
    
    for key, value in request.POST.items():
        print(f'Key: {key}')
        print(f'Value: {value}')
    
    #tools_list = Tool.objects.order_by('name')
    context = {
        #'tools_list': tools_list,
        'name': request.session.get("name")
    }

    # Open excel file on first arrival on the page (for each sheet)


    # Parse the Excel file for infos
    planning = settings.FILE_EXCEL_PLANNING
    wb = openpyxl.load_workbook(planning)
    ws = wb.active
    categories_list = []
    tools_list = []
    for row in range(4, ws.max_row):
        category = ws.cell(row=row, column=1).value
        if category is not None :
            categories_list.append(category)
        
        tool = ws.cell(row=row, column=2).value
        if (categories_list[-1] == request.POST.get("id_cat") 
            and tool is not None):
            tools_list.append(tool)
            
    context['categories_list'] = categories_list
    context['tools_list'] = tools_list
    wb.close()
    return render(request, 'main/sortie.html', context)

@login_required
def retour(request):
    return HttpResponse("Retour")

@login_required
def pret(request):
    return HttpResponse("Prêt")

@login_required
def recherche(request):
    context = {'name': request.session.get("name"),
    }
    
    # Check for AJAX POST (autopost every x seconds)
    # Show new tool name without reload
    if request.method == 'POST' :
        if 'autopost' in request.POST :
            return HttpResponse(SelectedTool.objects.all()[0].name)

    return render(request, 'main/recherche.html', context)

class Singleton(type):
    _instances = {}
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super(Singleton, cls).__call__(*args, **kwargs)
        return cls._instances[cls]
    
class VideoCamera(metaclass=Singleton):
    """
    Access the webcam of the computer.

    Args:
        metaclass : Singleton to keep one instance of camera on whole app
    """
    def __init__(self, request):
        print("init")
        self.request = request
        nb_webcam = settings.NB_WEBCAM_TO_USE
        self.video = cv2.VideoCapture(nb_webcam-1)
        (self.grabbed, self.frame) = self.video.read()
        t = threading.Thread(target=self.update, args=())
        t.start()
        
    def __del__(self):
        self.video.release()
        
    def get_frame(self):
        success, image = self.video.read()
        ret, jpeg = cv2.imencode('.jpg', image)
        return jpeg.tobytes()
            
    def update(self):
        print("update")
        while True :
            (self.grabbed, self.frame) = self.video.read()
            ### Only EAN13 to suppres warning about other types , symbols=[ZBarSymbol.EAN13 ]
            result = decode(self.frame)
            if result != []:
                for i in result:
                    print(i.data.decode('utf-8'))
                    st = SelectedTool.objects.all()[0]
                    st.name = i.data.decode('utf-8')
                    st.save()

def gen(camera):
    while True :
        frame = camera.get_frame()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')
        
def video_feed(request):
    return StreamingHttpResponse(gen(VideoCamera()),
                                     content_type='multipart/x-mixed-replace; boundary=frame')
    
def tool_feed(request):
    return HttpResponse(SelectedTool.objects.all()[0].name)
